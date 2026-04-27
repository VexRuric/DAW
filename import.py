#!/usr/bin/env python3
"""
One-time import: DAW Warehouse LIVE xlsx -> seed.sql for Supabase.

Usage:
    python import.py /path/to/DAW_Warehouse_LIVE_Roster_and_Stats.xlsx > seed.sql

Then copy seed.sql into Supabase SQL Editor and run.

NOTE: This is a best-effort import. Because the Google Form used free-text
participant fields, some matches will need manual review — especially:
  - misspelled wrestler names
  - "Abyss" and other WWE names that aren't in our roster
  - participant strings with unusual punctuation

The script writes a report of unresolved names to stderr. Review that list
before running seed.sql.
"""
import sys
import re
import uuid
from collections import defaultdict
from datetime import date, datetime
import openpyxl

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(s):
    """Strip whitespace and trailing newlines from free-text cells."""
    if s is None:
        return None
    return re.sub(r'\s+', ' ', str(s)).strip() or None

def sql_str(s):
    """Escape a value for SQL, or NULL."""
    if s is None or s == '':
        return 'NULL'
    if isinstance(s, bool):
        return 'TRUE' if s else 'FALSE'
    if isinstance(s, (int, float)):
        return str(s)
    if isinstance(s, (date, datetime)):
        return f"'{s.isoformat()[:10]}'"
    return "'" + str(s).replace("'", "''") + "'"

def norm_match_type(t):
    """Normalize 'Tag Team\\n' -> 'Tag Team' and consolidate aliases."""
    if not t:
        return 'Single'
    t = clean(t)
    aliases = {
        '5 Man': '5-Man', '5-Man': '5-Man',
        '6 Man': '6-Man',
        '6 Man Tag Team': '6-Man Tag Team',
        '3 Man Tag': '3-Man Tag',
        '3-Way Tornado Tag Team': 'Tornado Tag Team 3-Way',
        'Tag Team\n': 'Tag Team',
        'Title Match\n': 'Single',   # title match flag is separate
        'Triple Threat\nTag Team': 'Triple Threat Tag Team',
        'Royale Rumble': 'Royal Rumble',
        '1 vs 2': 'Handicap (1v2)',
        '1 vs 3': 'Handicap (1v3)',
    }
    return aliases.get(t, t)

def norm_defeat_type(d):
    """'Def.' / 'Def' / 'Def. (Pin)' -> 'Pin', 'Sub', 'DQ', 'No Contest'."""
    if not d:
        return None
    d = clean(d)
    if 'Pin' in d: return 'Pin'
    if 'Sub' in d: return 'Sub'
    if 'DQ' in d:  return 'DQ'
    if 'No Contest' in d: return 'No Contest'
    return 'Pin'   # bare "Def." defaults to pin

def scale_rating(r):
    """Sheet uses 1-10 (0.5 steps), we store 0.5-5.0."""
    if r is None or r == '':
        return None
    try:
        val = float(r)
    except (TypeError, ValueError):
        return None
    return round(val / 2, 1)

# Match type -> (min, max) participant count, so we can sanity check
PARTICIPANT_COUNT = {
    'Single': (2, 2),
    'Triple Threat': (3, 3),
    'Fatal 4-Way': (4, 4),
    '5-Man': (5, 5),
    '6-Man': (6, 6),
    'Battle Royale': (4, 30),
    'Royal Rumble': (10, 40),
}

# ---------------------------------------------------------------------------
# Participant string parsing
# ---------------------------------------------------------------------------

PAREN_RE = re.compile(r'\s*\([^)]*\)')

def strip_team_members(s):
    """'S&S Express (Slaghammers and Slapjax)' -> 'S&S Express'"""
    return PAREN_RE.sub('', s).strip()

def split_participants(raw):
    """
    '"Sweets vs Speczii vs Echo"' -> ['Sweets', 'Speczii', 'Echo']
    'S&S Express (Slaghammers and Slapjax) vs Buddy and Vac (...)'
      -> ['S&S Express', 'Buddy and Vac']
    Returns a list of stripped names (either wrestler or team).
    """
    if not raw:
        return []
    s = clean(raw)
    # Split on " vs " (with optional whitespace/newlines)
    parts = re.split(r'\s+vs\.?\s+|\s+v\.?\s+', s, flags=re.IGNORECASE)
    out = []
    for p in parts:
        name = strip_team_members(p).strip(' ,.').strip()
        if name:
            out.append(name)
    return out

def split_losers_smart(raw, known_names):
    """
    Losers can be 'A, B, C and D' format in multi-man matches.
    Because 'and' appears in team names ('Capn and the Willie', 'Buddy and Vac'),
    we first try to match known team/wrestler names greedily before splitting.
    """
    if not raw or clean(raw) in ('None', 'none', 'N/A', 'NONE'):
        return []
    s = clean(raw)
    s = strip_team_members(s)  # remove '(members)' groups first

    # Greedy match: try to peel known names off the front of the string
    known_sorted = sorted(known_names, key=len, reverse=True)
    remaining = s
    found = []
    while remaining:
        remaining = remaining.strip(' ,.&')
        if not remaining:
            break
        matched = None
        for name in known_sorted:
            # Match if remaining starts with this name (case-insensitive)
            if remaining.lower().startswith(name.lower()):
                # Make sure we matched a whole word boundary
                nxt = remaining[len(name):len(name)+1]
                if nxt in ('', ' ', ',', '.', '&'):
                    matched = name
                    break
        if matched:
            found.append(matched)
            remaining = remaining[len(matched):]
        else:
            # Couldn't match — try splitting off the next comma/and chunk
            m = re.search(r',\s*|\s+and\s+|\s+&\s+', remaining)
            if m:
                chunk = remaining[:m.start()].strip()
                remaining = remaining[m.end():]
                if chunk:
                    found.append(chunk)
            else:
                if remaining.strip():
                    found.append(remaining.strip())
                break
    return found

# ---------------------------------------------------------------------------
# Title name normalization
# ---------------------------------------------------------------------------

TITLE_CANON = {
    'Mens Champion':             ('Mens Champion', 'Singles', 'Mens'),
    'Womens Champion':           ('Womens Champion', 'Singles', 'Womens'),
    'Intergender Champion':      ('Intergender Champion', 'Singles', 'Any'),
    'Internet Champion':         ('Internet Champion', 'Singles', 'Any'),
    'Livesteam Champion':        ('Internet Champion', 'Singles', 'Any'),  # old name
    'Intercontinental Champion': ('Intercontinental Champion', 'Singles', 'Any'),
    'Hardcore Champion':         ('Hardcore Champion', 'Singles', 'Any'),
    'Mens Chat Champion':        ('Mens Chat Champion', 'Singles', 'Mens'),
    'Womens Chat Champion':      ('Womens Chat Champion', 'Singles', 'Womens'),
    'Mens Tag Team Champion':    ('Mens Tag Team Champion', 'Tag', 'Mens'),
    'Womens Tag Team Champion':  ('Womens Tag Team Champion', 'Tag', 'Womens'),
    'World Tag Team Champion':   ('World Tag Team Champion', 'Tag', 'Any'),
    'Mens Money in the Bank':    ("Mr Money in the Bank", 'MitB', 'Mens'),
    'Womens Money in the Bank':  ("Miss Money in the Bank", 'MitB', 'Womens'),
    'Mens Tag Money in the Bank':("Mens Tag MitB", 'MitB', 'Mens'),
    'MitB':                      ("Mr Money in the Bank", 'MitB', 'Mens'),
    'Cross Brand Mens Champion': ('Cross Brand Mens Champion', 'Singles', 'Mens'),
    'Cross Brand Womens Champion':('Cross Brand Womens Champion', 'Singles', 'Womens'),
    'Cross Brand Tag Champion':  ('Cross Brand Tag Champion', 'Tag', 'Any'),
    'Mayor':                     ('Mayor', 'Special', 'Any'),
    'Mayor Title':               ('Mayor', 'Special', 'Any'),
}

STIPULATION_CANON = {
    'Extreme Rules\n': 'Extreme Rules',
    'Ladder\n': 'Ladder',
    'Hell in a Cell Match': 'Hell in a Cell',
    'Table Ladder Chairs': 'TLC',
}

def canonical_title(raw):
    if not raw:
        return None
    c = clean(raw)
    return TITLE_CANON.get(c, (c, 'Singles', 'Any'))[0] if c in TITLE_CANON else None

def canonical_stipulation(raw):
    if not raw:
        return None
    c = clean(raw)
    # If it's actually a title name, that's not a stipulation
    if c in TITLE_CANON:
        return None
    return STIPULATION_CANON.get(c, c)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---------- Wrestlers ----------
    ws = wb['Roster']
    wrestlers = {}   # name -> dict
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        if not name:
            continue
        wrestlers[name.lower()] = {
            'id': str(uuid.uuid4()),
            'name': name,
            'brand': clean(row[2]),
            'division': clean(row[3]),
            'active': clean(row[4]) == 'Yes',
            'role': clean(row[5]),
            'gender': clean(row[6]),
            'gimmick': clean(row[7]),
            'injured': clean(row[9]) == 'Yes',
            'twitch_handle': clean(row[10]),
            'disposition': row[11] if isinstance(row[11], (int, float)) else None,
        }

    # ---------- Teams ----------
    ws = wb['Team Roster']
    teams = {}    # name -> dict
    team_members = []  # list of (team_name, wrestler_name)
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        if not name:
            continue
        teams[name.lower()] = {
            'id': str(uuid.uuid4()),
            'name': name,
            'brand': clean(row[2]),
            'division': clean(row[3]),
            'active': clean(row[4]) == 'Yes',
            'role': clean(row[5]),
            'gender': clean(row[6]),
            'injured': clean(row[11]) == 'Yes',
        }
        for member in (row[7], row[8], row[9], row[10]):
            m = clean(member)
            if m:
                team_members.append((name.lower(), m.lower()))

    # ---------- Titles (derived from canonical map) ----------
    titles = {}
    for canonical_name, (n, cat, gender) in {v: v for v in TITLE_CANON.values()}.items():
        pass
    seen = set()
    for _, (cname, cat, gender) in TITLE_CANON.items():
        if cname in seen:
            continue
        seen.add(cname)
        titles[cname.lower()] = {
            'id': str(uuid.uuid4()),
            'name': cname,
            'category': cat,
            'gender': gender,
            'active': 'Cross Brand' not in cname and 'Mayor' not in cname,
        }

    # ---------- Matches ----------
    ws = wb['Form Responses 4']
    matches_by_show = defaultdict(list)   # (date, show_type) -> list of match dicts
    unresolved = defaultdict(int)
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:   # no match date
            skipped += 1
            continue
        mdate = row[1]
        if isinstance(mdate, datetime):
            mdate = mdate.date()
        mnum = int(row[2]) if isinstance(row[2], (int, float)) else None
        if mnum is None:
            skipped += 1
            continue

        mtype = norm_match_type(row[3])
        participants_raw = row[4]
        winner_raw = row[5]
        defeat = norm_defeat_type(row[6])
        losers_raw = row[7]
        is_draw = bool(clean(row[8]))
        rating = scale_rating(row[9])
        is_title = clean(row[10]) == 'Yes'
        title_or_mitb_raw = clean(row[11])
        stip_raw = clean(row[12])

        # Figure out which field holds the title and which holds the stip
        title_name = canonical_title(title_or_mitb_raw) if is_title else None
        stipulation = canonical_stipulation(stip_raw)
        if not stipulation:
            stipulation = canonical_stipulation(title_or_mitb_raw) if not is_title else None

        is_mitb = (mtype == 'MitB') or (title_or_mitb_raw and 'Money in the Bank' in (title_or_mitb_raw or ''))
        cashin = (title_or_mitb_raw == 'Mr MitB Cash-In')

        participants = split_participants(participants_raw)
        winners = split_participants(winner_raw) if winner_raw else []
        known_names = list(wrestlers.keys()) + list(teams.keys())
        # Build pretty-cased lookup for matching
        pretty = {}
        for w in wrestlers.values(): pretty[w['name'].lower()] = w['name']
        for t in teams.values(): pretty[t['name'].lower()] = t['name']
        known_pretty = list(pretty.values())
        losers = split_losers_smart(losers_raw, known_pretty) if losers_raw else []

        # Resolve each name to wrestler or team
        def resolve(name):
            low = name.lower()
            if low in teams:
                return ('team', teams[low]['id'])
            if low in wrestlers:
                return ('wrestler', wrestlers[low]['id'])
            unresolved[name] += 1
            return None

        resolved_parts = [(n, resolve(n)) for n in participants]
        resolved_winners = [(n, resolve(n)) for n in winners]
        resolved_losers = [(n, resolve(n)) for n in losers]

        # Show key: a single "show" per date (weekly) or per PPV name
        show_key = (mdate, 'weekly', None)  # TODO: detect PPV by date

        matches_by_show[show_key].append({
            'match_number': mnum,
            'match_type': mtype,
            'stipulation': stipulation,
            'is_title_match': is_title,
            'title_name': title_name,
            'is_mitb': is_mitb,
            'mitb_cashin': cashin,
            'defeat_type': defeat,
            'rating': rating,
            'is_draw': is_draw,
            'participants': resolved_parts,
            'winners': resolved_winners,
            'losers': resolved_losers,
        })

    # ---------- Write SQL ----------
    out = sys.stdout
    out.write("-- Generated by import.py\n")
    out.write("begin;\n\n")

    # Wrestlers
    out.write("-- Wrestlers\n")
    for w in wrestlers.values():
        out.write(
            f"insert into wrestlers (id, name, brand, division, active, role, gender, "
            f"gimmick, injured, twitch_handle, disposition) values "
            f"({sql_str(w['id'])}, {sql_str(w['name'])}, {sql_str(w['brand'])}, "
            f"{sql_str(w['division'])}, {sql_str(w['active'])}, {sql_str(w['role'])}, "
            f"{sql_str(w['gender'])}, {sql_str(w['gimmick'])}, {sql_str(w['injured'])}, "
            f"{sql_str(w['twitch_handle'])}, {sql_str(w['disposition'])});\n"
        )
    out.write("\n")

    # Teams
    out.write("-- Teams\n")
    for t in teams.values():
        out.write(
            f"insert into teams (id, name, brand, division, active, role, gender, injured) values "
            f"({sql_str(t['id'])}, {sql_str(t['name'])}, {sql_str(t['brand'])}, "
            f"{sql_str(t['division'])}, {sql_str(t['active'])}, {sql_str(t['role'])}, "
            f"{sql_str(t['gender'])}, {sql_str(t['injured'])});\n"
        )
    out.write("\n")

    # Team memberships
    out.write("-- Team memberships\n")
    for team_name_low, wrestler_name_low in team_members:
        t = teams.get(team_name_low)
        w = wrestlers.get(wrestler_name_low)
        if t and w:
            out.write(
                f"insert into team_memberships (team_id, wrestler_id) values "
                f"({sql_str(t['id'])}, {sql_str(w['id'])});\n"
            )
    out.write("\n")

    # Titles
    out.write("-- Titles\n")
    for i, title in enumerate(titles.values()):
        out.write(
            f"insert into titles (id, name, category, gender, active, display_order) values "
            f"({sql_str(title['id'])}, {sql_str(title['name'])}, {sql_str(title['category'])}, "
            f"{sql_str(title['gender'])}, {sql_str(title['active'])}, {i});\n"
        )
    out.write("\n")

    # Shows + matches + participants
    out.write("-- Shows and matches\n")
    for (show_date, show_type, ppv), match_list in sorted(matches_by_show.items()):
        show_id = str(uuid.uuid4())
        show_name = f"DAW {show_date.strftime('%m-%d-%Y')}"
        out.write(
            f"insert into shows (id, name, show_date, show_type, ppv_name) values "
            f"({sql_str(show_id)}, {sql_str(show_name)}, {sql_str(show_date)}, "
            f"{sql_str(show_type)}, {sql_str(ppv)});\n"
        )
        for m in sorted(match_list, key=lambda x: x['match_number']):
            match_id = str(uuid.uuid4())
            title_id = titles.get((m['title_name'] or '').lower(), {}).get('id') if m['title_name'] else None
            out.write(
                f"insert into matches (id, show_id, match_number, match_type, stipulation, "
                f"is_title_match, title_id, is_mitb, mitb_cashin, defeat_type, rating, is_draw) values "
                f"({sql_str(match_id)}, {sql_str(show_id)}, {m['match_number']}, "
                f"{sql_str(m['match_type'])}, {sql_str(m['stipulation'])}, "
                f"{sql_str(m['is_title_match'])}, {sql_str(title_id)}, "
                f"{sql_str(m['is_mitb'])}, {sql_str(m['mitb_cashin'])}, "
                f"{sql_str(m['defeat_type'])}, {sql_str(m['rating'])}, "
                f"{sql_str(m['is_draw'])});\n"
            )
            # participants
            winner_ids = {ref[1] for _, ref in m['winners'] if ref}
            loser_ids = {ref[1] for _, ref in m['losers'] if ref}
            for name, ref in m['participants']:
                if not ref:
                    continue
                kind, rid = ref
                if rid in winner_ids:
                    result = 'winner'
                elif rid in loser_ids:
                    result = 'loser'
                elif m['is_draw']:
                    result = 'draw'
                else:
                    result = 'loser'  # default: if not winner, loser
                col = 'wrestler_id' if kind == 'wrestler' else 'team_id'
                out.write(
                    f"insert into match_participants (match_id, {col}, result) values "
                    f"({sql_str(match_id)}, {sql_str(rid)}, {sql_str(result)});\n"
                )

    out.write("\ncommit;\n")

    # Report unresolved names to stderr
    if unresolved:
        sys.stderr.write("\n=== UNRESOLVED NAMES (review before running seed.sql) ===\n")
        for name, count in sorted(unresolved.items(), key=lambda x: -x[1])[:30]:
            sys.stderr.write(f"  {name!r}  (appears {count} times)\n")
        sys.stderr.write(f"  ... and {max(0, len(unresolved)-30)} more. Full list written to cleanup.csv\n")

        # Write cleanup CSV with suggested mappings
        import csv
        with open('cleanup.csv', 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['unresolved_name', 'occurrences', 'suggested_match', 'action_needed'])
            all_names = sorted(set(list(wrestlers.keys()) + list(teams.keys())))
            for name, count in sorted(unresolved.items(), key=lambda x: -x[1]):
                # Suggest a match using simple similarity
                suggestion = ''
                low = name.lower().strip(' ,.)(')
                for known in all_names:
                    if known.startswith(low[:4]) or low.startswith(known[:4]):
                        suggestion = known
                        break
                action = 'add_as_new' if count >= 3 else 'review'
                w.writerow([name, count, suggestion, action])

    sys.stderr.write(f"\nSkipped {skipped} malformed rows.\n")
    sys.stderr.write(f"Imported {sum(len(v) for v in matches_by_show.values())} matches "
                     f"across {len(matches_by_show)} shows.\n")
    sys.stderr.write(f"Wrestlers: {len(wrestlers)}, Teams: {len(teams)}, Titles: {len(titles)}\n")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.stderr.write("Usage: python import.py <xlsx_path>\n")
        sys.exit(1)
    main(sys.argv[1])
